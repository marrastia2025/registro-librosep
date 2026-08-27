import streamlit as st
import sqlite3
import easyocr
import numpy as np
import pandas as pd
import io
from PIL import Image as PILImage
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage

# ------------------- CONFIGURACIÓN -------------------
archivo_excel = "libros.xlsx"
reader = easyocr.Reader(['es'])

# ------------------- BASE DE DATOS -------------------
def init_db():
    conn = sqlite3.connect("usuarios.db")
    cursor = conn.cursor()
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS usuarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        usuario TEXT UNIQUE,
        password TEXT,
        rol TEXT
    )
    """)
    # Crear admin por defecto si no existe
    cursor.execute("INSERT OR IGNORE INTO usuarios (usuario, password, rol) VALUES ('admin', '1234', 'admin')")
    conn.commit()
    conn.close()

def validar_usuario(usuario, password):
    conn = sqlite3.connect("usuarios.db")
    cursor = conn.cursor()
    cursor.execute("SELECT rol FROM usuarios WHERE usuario=? AND password=?", (usuario, password))
    resultado = cursor.fetchone()
    conn.close()
    if resultado:
        return resultado[0]
    return None

init_db()

# ------------------- FUNCIONES LIBROS -------------------
def crear_archivo():
    try:
        wb = openpyxl.load_workbook(archivo_excel)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Libros"
        ws.append(["Título", "Autor", "Editorial", "Foto Portada"])
        wb.save(archivo_excel)

crear_archivo()

def listar_libros():
    wb = openpyxl.load_workbook(archivo_excel)
    ws = wb["Libros"]
    data = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        data.append(fila)
    df = pd.DataFrame(data, columns=["Título", "Autor", "Editorial", "Foto Portada"])
    return df

def registrar_libro(titulo, autor, editorial, foto):
    wb = openpyxl.load_workbook(archivo_excel)
    ws = wb["Libros"]
    fila = ws.max_row + 1
    ws.cell(row=fila, column=1, value=titulo)
    ws.cell(row=fila, column=2, value=autor)
    ws.cell(row=fila, column=3, value=editorial)

    img_bytes = io.BytesIO(foto.read())
    img_temp = PILImage.open(img_bytes)
    img_temp.save("temp_portada.png")
    img_excel = ExcelImage("temp_portada.png")
    img_excel.width, img_excel.height = 100, 150
    ws.add_image(img_excel, f"D{fila}")

    wb.save(archivo_excel)

def actualizar_portada(titulo, foto):
    wb = openpyxl.load_workbook(archivo_excel)
    ws = wb["Libros"]
    for fila in range(2, ws.max_row + 1):
        if ws.cell(row=fila, column=1).value == titulo:
            img_bytes = io.BytesIO(foto.read())
            img_temp = PILImage.open(img_bytes)
            img_temp.save("temp_portada.png")
            img_excel = ExcelImage("temp_portada.png")
            img_excel.width, img_excel.height = 100, 150
            ws.add_image(img_excel, f"D{fila}")
            wb.save(archivo_excel)
            return True
    return False

def eliminar_libro(titulo):
    wb = openpyxl.load_workbook(archivo_excel)
    ws = wb["Libros"]
    for fila in range(2, ws.max_row + 1):
        if ws.cell(row=fila, column=1).value == titulo:
            ws.delete_rows(fila)
            wb.save(archivo_excel)
            return True
    return False

# ------------------- LOGIN -------------------
st.title("📚 Registro de Libros con OCR (EasyOCR)")

if "logueado" not in st.session_state:
    st.session_state.logueado = False
    st.session_state.rol = None

if not st.session_state.logueado:
    st.subheader("🔐 Iniciar sesión")
    usuario = st.text_input("Usuario")
    clave = st.text_input("Contraseña", type="password")
    if st.button("Login"):
        rol = validar_usuario(usuario, clave)
        if rol:
            st.session_state.logueado = True
            st.session_state.rol = rol
            st.success(f"✅ Acceso concedido como {rol}")
        else:
            st.error("❌ Usuario o contraseña incorrectos")
else:
    rol = st.session_state.rol
    st.info(f"👤 Rol actual: {rol}")

    # Botón de logout
    if st.button("Logout"):
        st.session_state.logueado = False
        st.session_state.rol = None
        st.experimental_rerun()

    # ------------------- APP PRINCIPAL -------------------
    if rol == "admin":
        foto = st.file_uploader("📷 Subir portada del libro", type=["jpg","jpeg","png"])
        if foto:
            st.image(foto, caption="Portada subida", use_column_width=True)

            # OCR con EasyOCR
            imagen = PILImage.open(foto)
            texto_detectado = reader.readtext(np.array(imagen), detail=0)
            texto = "\n".join(texto_detectado)
            st.text_area("Texto detectado", texto)

            lineas = [l.strip() for l in texto_detectado if l.strip()]
            titulo = st.text_input("Título", lineas[0] if len(lineas) > 0 else "")
            autor = st.text_input("Autor", lineas[1] if len(lineas) > 1 else "")
            editorial = st.text_input("Editorial", lineas[2] if len(lineas) > 2 else "")

            if st.button("Guardar Registro"):
                registrar_libro(titulo, autor, editorial, foto)
                st.success("✅ Libro registrado correctamente con imagen en Excel.")

    # ------------------- LISTADO -------------------
    st.subheader("📊 Listado de Libros Registrados")
    df_libros = listar_libros()
    for i, row in df_libros.iterrows():
        col1, col2, col3, col4, col5 = st.columns([3, 3, 3, 1, 2])
        col1.write(row["Título"])
        col2.write(row["Autor"])
        col3.write(row["Editorial"])

        if rol == "admin":
            if col4.button("🗑️ Eliminar", key=f"del_{i}"):
                if eliminar_libro(row["Título"]):
                    st.success(f"Libro '{row['Título']}' eliminado correctamente.")
                    st.experimental_rerun()

            nueva_portada = col5.file_uploader(f"Configurar portada {row['Título']}", type=["jpg","jpeg","png"], key=f"portada_{i}")
            if nueva_portada and col5.button("Actualizar", key=f"upd_{i}"):
                if actualizar_portada(row["Título"], nueva_portada):
                    st.success(f"✅ Portada de '{row['Título']}' actualizada.")
                    st.experimental_rerun()
